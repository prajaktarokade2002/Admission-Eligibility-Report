import tkinter as tk
from tkinter import filedialog, messagebox
import pdfplumber
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
#from openpyxl.utils.dataframe import dataframe_to_rows, get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# === PDF Processing Functions ===

def process_fy_to_sy(pdf_path, academic_year, student_class, exam_period, fy_min, fy_max):
    records = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            name_match = re.search(r"Name\s*:\s*([^\n\r]+?)(?:\s*Mother Name:|$)", text)
            credit_match = re.search(r"Total\s+Credits\s*:\s*(\d+)", text)
            if name_match and credit_match:
                name = name_match.group(1).strip()
                credits = int(credit_match.group(1))
                eligibility = "Eligible" if fy_min <= credits <= fy_max else "Not Eligible"
                records.append({
                    "Student Name": name,
                    "Total Credits": credits,
                    "Status": eligibility,
                    "Exam Period": exam_period
                })
    return pd.DataFrame(records)

def process_sy_to_ty(pdf_path, academic_year, student_class, exam_period, fy_min, fy_max, sy_min, sy_max):
    results = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            student_match = re.search(r"Student's Name[:\-]?\s*(.+)", text)
            student_name = student_match.group(1).strip() if student_match else "Unknown"

            fy_match = re.search(r"FY\s*/\s*Part\s*1:\s*(.*)", text)
            sy_match = re.search(r"SY\s*/\s*Part\s*2:\s*(.*)", text)

            def extract_total(line):
                parts = line.split()
                if len(parts) >= 3:
                    return parts[2]
                return None

            fy_total_str = extract_total(fy_match.group(1)) if fy_match else None
            sy_total_str = extract_total(sy_match.group(1)) if sy_match else None

            try:
                fy_total = int(fy_total_str) if fy_total_str and fy_total_str.isdigit() else None
                sy_total = int(sy_total_str) if sy_total_str and sy_total_str.isdigit() else None
            except ValueError:
                fy_total = sy_total = None

            status = "Eligible" if (fy_total is not None and fy_min <= fy_total <= fy_max) and (sy_total is not None and sy_min <= sy_total <= sy_max) else "Not Eligible"

            results.append({
                "Student Name": student_name,
                "FY Total": fy_total,
                "SY Total": sy_total,
                "Status": status,
                "Exam Period": exam_period
            })

    return pd.DataFrame(results)

# === Excel Report Generator ===

def generate_excel_report(df, academic_year, student_class, exam_period, output_path):
    wb = Workbook()
    ws = wb.active
    last_col = get_column_letter(len(df.columns))

    # Headings
    def create_heading(row, text, size=12):
        cell = f"A{row}"
        ws[cell] = text
        ws.merge_cells(f"A{row}:{last_col}{row}")
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[cell].font = Font(bold=True, size=size)

    headings = [
        "Progressive Education Society's",
        "Modern College of Arts, Science and Commerce",
        "Ganeshkhind, Pune-16",
        f"Class: {student_class}",
        f"Academic Year: {academic_year}",
        f"Exam Period: {exam_period}"
    ]

    for i, line in enumerate(headings, 1):
        create_heading(i, line)

    # Data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=7):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Column width
    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=7, max_row=7, max_col=len(df.columns)), start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Borders
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=7, max_row=7 + len(df), min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.border = thin_border

    wb.save(output_path)

# === GUI Handlers ===

def browse_pdf():
    filename = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    pdf_entry.delete(0, tk.END)
    pdf_entry.insert(0, filename)

def browse_output_folder():
    foldername = filedialog.askdirectory()
    output_entry.delete(0, tk.END)
    output_entry.insert(0, foldername)

def run_extraction():
    pdf_path = pdf_entry.get().strip()
    academic_year = year_entry.get().strip()
    student_class = class_entry.get().strip()
    exam_period = exam_period_var.get().strip()
    output_dir = output_entry.get().strip()
    eligibility_type = eligibility_type_var.get()

    if not os.path.isfile(pdf_path):
        messagebox.showerror("Error", "Please select a valid PDF file.")
        return
    if not academic_year or not student_class or not exam_period:
        messagebox.showerror("Error", "Please fill in Academic Year, Class, and Exam Period.")
        return
    if not os.path.isdir(output_dir):
        messagebox.showerror("Error", "Please select a valid output folder.")
        return

    try:
        if eligibility_type == "fy_sy":
            df = process_fy_to_sy(pdf_path, academic_year, student_class, exam_period,
                                  fy_min_var.get(), fy_max_var.get())
        else:
            df = process_sy_to_ty(pdf_path, academic_year, student_class, exam_period,
                                  fy_min_var.get(), fy_max_var.get(),
                                  sy_min_var.get(), sy_max_var.get())

        filename = f"{student_class}_{eligibility_type.upper()}_Report_{academic_year.replace('-', '_')}.xlsx"
        output_file = os.path.join(output_dir, filename)

        generate_excel_report(df, academic_year, student_class, exam_period, output_file)
        messagebox.showinfo("Success", f"Report saved to:\n{output_file}")
    except Exception as e:
        messagebox.showerror("Error", str(e))

# === GUI Setup ===

root = tk.Tk()
root.title("Student Eligibility Report Generator")

# File selection
tk.Label(root, text="Select PDF File:").grid(row=0, column=0, sticky="e", padx=10, pady=5)
pdf_entry = tk.Entry(root, width=50)
pdf_entry.grid(row=0, column=1, padx=10)
tk.Button(root, text="Browse", command=browse_pdf).grid(row=0, column=2, padx=5)

# Academic Year
tk.Label(root, text="Academic Year (YYYY-YYYY):").grid(row=1, column=0, sticky="e", padx=10, pady=5)
year_entry = tk.Entry(root, width=30)
year_entry.grid(row=1, column=1, padx=10)

# Class
tk.Label(root, text="Class:").grid(row=2, column=0, sticky="e", padx=10, pady=5)
class_entry = tk.Entry(root, width=30)
class_entry.grid(row=2, column=1, padx=10)

# Exam Period
tk.Label(root, text="Exam Period:").grid(row=3, column=0, sticky="e", padx=10, pady=5)
exam_period_var = tk.StringVar(value="Mar-April")
tk.Radiobutton(root, text="Mar-April", variable=exam_period_var, value="Mar-April").grid(row=3, column=1, sticky="w")
tk.Radiobutton(root, text="Oct-Nov", variable=exam_period_var, value="Oct-Nov").grid(row=4, column=1, sticky="w")

# Eligibility Type (horizontal layout)
tk.Label(root, text="Eligibility Type:").grid(row=5, column=0, sticky="e", padx=10, pady=5)
eligibility_type_var = tk.StringVar(value="fy_sy")
frame = tk.Frame(root)
frame.grid(row=5, column=1, sticky="w")
tk.Radiobutton(frame, text="FY to SY", variable=eligibility_type_var, value="fy_sy").pack(side="left", padx=5)
tk.Radiobutton(frame, text="SY to TY", variable=eligibility_type_var, value="sy_ty").pack(side="left", padx=5)

# === Conditional Rules Section ===
condition_frame = tk.Frame(root)
condition_frame.grid(row=6, column=1, columnspan=2, sticky="w", pady=5)

# Default values (user can edit)
fy_min_var = tk.IntVar(value=22)
fy_max_var = tk.IntVar(value=44)
sy_min_var = tk.IntVar(value=22)
sy_max_var = tk.IntVar(value=44)

def update_condition_fields(*args):
    for widget in condition_frame.winfo_children():
        widget.destroy()

    selected_type = eligibility_type_var.get()

    if selected_type == "fy_sy":
        tk.Label(condition_frame, text="FY Credits Range:").pack(side="left", padx=5)
        tk.Entry(condition_frame, textvariable=fy_min_var, width=5).pack(side="left")
        tk.Label(condition_frame, text="to").pack(side="left")
        tk.Entry(condition_frame, textvariable=fy_max_var, width=5).pack(side="left")

    elif selected_type == "sy_ty":
        tk.Label(condition_frame, text="FY Credits Range:").pack(side="left", padx=5)
        tk.Entry(condition_frame, textvariable=fy_min_var, width=5).pack(side="left")
        tk.Label(condition_frame, text="to").pack(side="left")
        tk.Entry(condition_frame, textvariable=fy_max_var, width=5).pack(side="left")

        tk.Label(condition_frame, text="   SY Credits Range:").pack(side="left", padx=5)
        tk.Entry(condition_frame, textvariable=sy_min_var, width=5).pack(side="left")
        tk.Label(condition_frame, text="to").pack(side="left")
        tk.Entry(condition_frame, textvariable=sy_max_var, width=5).pack(side="left")

eligibility_type_var.trace_add("write", update_condition_fields)
update_condition_fields()

# Output Folder
tk.Label(root, text="Select Output Folder:").grid(row=7, column=0, sticky="e", padx=10, pady=5)
output_entry = tk.Entry(root, width=50)
output_entry.grid(row=7, column=1, padx=10)
tk.Button(root, text="Browse", command=browse_output_folder).grid(row=7, column=2, padx=5)

# Extract Button
tk.Button(root, text="Generate Excel Report", command=run_extraction, bg="green", fg="white").grid(row=8, column=1, pady=20)

root.mainloop()
